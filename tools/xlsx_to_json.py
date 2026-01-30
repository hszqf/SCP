#!/usr/bin/env python3
"""Export game_data.xlsx into Unity-friendly game_data.json."""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

LIST_SPLIT_PATTERN = re.compile(r"[;,，]")
ALLOWED_TYPES = {"int", "float", "string", "int[]", "float[]", "string[]"}
LOG_LEVELS = {"DEBUG": 10, "INFO": 20, "WARN": 30, "ERROR": 40}


@dataclass(slots=True)
class ColumnInfo:
    index: int
    name: str
    type_name: str


@dataclass(slots=True)
class SheetTable:
    name: str
    id_field: str
    columns: list[dict[str, str]]
    rows: list[dict[str, Any]]


class Reporter:
    def __init__(self, level_name: str) -> None:
        level = LOG_LEVELS.get(level_name.upper())
        if level is None:
            raise ValueError(f"Unsupported log level: {level_name}")
        self.level = level

    def _emit(self, label: str, message: str, *, error: bool = False) -> None:
        stream = sys.stderr if error else sys.stdout
        print(f"[{label}] {message}", file=stream)

    def info(self, message: str, *args: Any) -> None:
        if self.level <= LOG_LEVELS["INFO"]:
            self._emit("INFO", message % args if args else message)

    def warn(self, message: str, *args: Any) -> None:
        if self.level <= LOG_LEVELS["WARN"]:
            self._emit("WARN", message % args if args else message)

    def error(self, message: str, *args: Any) -> None:
        self._emit("ERROR", message % args if args else message, error=True)

    def success(self, message: str, *args: Any) -> None:
        self._emit("SUCCESS", message % args if args else message)


def split_list(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = [part.strip() for part in LIST_SPLIT_PATTERN.split(text)]
    return [part for part in parts if part]


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    return int(value)


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    return float(value)


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_cell(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _parse_table_value(type_name: str, value: Any) -> Any:
    if type_name == "string":
        return "" if value is None else str(value).strip()
    if type_name == "int":
        return to_int(value)
    if type_name == "float":
        return to_float(value)
    if type_name == "string[]":
        return split_list(value)
    if type_name == "int[]":
        return [int(part) for part in split_list(value)]
    if type_name == "float[]":
        return [float(part) for part in split_list(value)]
    raise ValueError(f"Unsupported type: {type_name}")


def parse_sheet(ws, reporter: Reporter) -> tuple[SheetTable | None, list[str]]:
    issues: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        reporter.warn("sheet %s empty", ws.title)
        return None, issues
    if len(rows) < 3:
        issues.append(f"{ws.title} requires at least 3 header rows")
        return None, issues

    header_row = rows[1]
    type_row = rows[2]
    data_rows = rows[3:]

    columns: list[dict[str, str]] = []
    column_infos: list[ColumnInfo] = []
    for idx, raw_name in enumerate(header_row):
        name = _normalize_header(raw_name)
        if not name or name.startswith("#"):
            continue
        raw_type = type_row[idx] if idx < len(type_row) else None
        type_name = _normalize_header(raw_type)
        if not type_name:
            issues.append(f"{ws.title} column {name} missing type")
            continue
        if type_name not in ALLOWED_TYPES:
            issues.append(f"{ws.title} column {name} invalid type {type_name!r}")
            continue
        columns.append({"name": name, "type": type_name})
        column_infos.append(ColumnInfo(index=idx, name=name, type_name=type_name))

    if not column_infos:
        issues.append(f"{ws.title} has no exportable columns")
        return None, issues

    id_field = column_infos[0].name
    id_rows: dict[Any, list[int]] = {}
    row_entries: list[dict[str, Any]] = []

    for row_offset, row in enumerate(data_rows):
        row_number = row_offset + 4
        if row is None:
            continue
        if all(_is_empty_cell(row[col.index] if col.index < len(row) else None) for col in column_infos):
            continue
        entry: dict[str, Any] = {}
        for col in column_infos:
            cell_value = row[col.index] if col.index < len(row) else None
            try:
                entry[col.name] = _parse_table_value(col.type_name, cell_value)
            except (TypeError, ValueError) as exc:
                issues.append(f"{ws.title}[row {row_number}].{col.name} parse error: {exc}")
        raw_id_cell = row[column_infos[0].index] if column_infos[0].index < len(row) else None
        if _is_empty_cell(raw_id_cell):
            issues.append(f"{ws.title}[row {row_number}] {id_field} is empty")
            continue
        id_value = entry.get(id_field)
        id_rows.setdefault(id_value, []).append(row_number)
        row_entries.append(entry)

    for id_value, rows_list in id_rows.items():
        if len(rows_list) > 1:
            issues.append(f"{ws.title} duplicate id {id_value!r} rows {rows_list}")

    reporter.info(
        "sheet %s cols=%d rows=%d exportedRows=%d",
        ws.title,
        len(column_infos),
        len(data_rows),
        len(row_entries),
    )
    return SheetTable(name=ws.title, id_field=id_field, columns=columns, rows=row_entries), issues


def build_tables(workbook, reporter: Reporter) -> tuple[dict[str, Any], list[str]]:
    tables: dict[str, Any] = {}
    issues: list[str] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        table, sheet_issues = parse_sheet(ws, reporter)
        issues.extend(sheet_issues)
        if table is None:
            continue
        tables[sheet_name] = {
            "idField": table.id_field,
            "columns": table.columns,
            "rows": table.rows,
        }
    return tables, issues


def find_git_root(start: Path) -> Path | None:
    try:
        completed = subprocess.run(
            ["git", "rev-parse", "--show-toplevel"],
            cwd=start,
            check=True,
            capture_output=True,
            text=True,
        )
    except (OSError, subprocess.CalledProcessError):
        return None
    root = completed.stdout.strip()
    return Path(root) if root else None


def resolve_project_root(explicit_root: str | None) -> Path:
    if explicit_root:
        return Path(explicit_root).expanduser().resolve()
    script_root = find_git_root(Path(__file__).resolve().parent)
    if script_root:
        return script_root.resolve()
    return Path.cwd().resolve()


def resolve_path(root: Path, value: str | None, default_rel: str) -> Path:
    raw = value or default_rel
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = root / path
    return path.resolve()


def parse_args(argv: list[str]) -> argparse.Namespace:
    legacy_xlsx: str | None = None
    legacy_out: str | None = None
    legacy_tail = argv[1:]
    if len(legacy_tail) >= 2 and not legacy_tail[0].startswith("-") and not legacy_tail[1].startswith("-"):
        legacy_xlsx, legacy_out = legacy_tail[0], legacy_tail[1]

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("legacy_xlsx", nargs="?", help=argparse.SUPPRESS)
    parser.add_argument("legacy_out", nargs="?", help=argparse.SUPPRESS)
    parser.add_argument("--xlsx", dest="xlsx", help="Path to game_data.xlsx")
    parser.add_argument("--out", dest="out", help="Path to output game_data.json")
    parser.add_argument(
        "--project-root",
        dest="project_root",
        help="Repository root; defaults to git root or current working directory.",
    )
    parser.add_argument(
        "--log-level",
        dest="log_level",
        default="INFO",
        choices=sorted({"DEBUG", "INFO", "WARN", "ERROR"}),
        help="Logging level (default: INFO)",
    )
    parser.add_argument(
        "--validate-only",
        "--dry-run",
        dest="validate_only",
        action="store_true",
        help="Validate workbook without writing output JSON.",
    )
    parser.add_argument(
        "--no-pretty",
        dest="no_pretty",
        action="store_true",
        help="Emit compact JSON instead of pretty-printed JSON.",
    )

    args = parser.parse_args(argv[1:])

    if not args.xlsx:
        args.xlsx = legacy_xlsx or args.legacy_xlsx
    if not args.out:
        args.out = legacy_out or args.legacy_out

    return args


def run(argv: list[str]) -> int:
    args = parse_args(argv)
    try:
        reporter = Reporter(args.log_level)
    except ValueError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    start_time = time.perf_counter()

    project_root = resolve_project_root(args.project_root)
    xlsx_path = resolve_path(project_root, args.xlsx, "GameData/Local/game_data.xlsx")
    out_path = resolve_path(project_root, args.out, "Assets/StreamingAssets/game_data.json")

    reporter.info("xlsx=%s out=%s", xlsx_path, out_path)

    if not xlsx_path.exists():
        reporter.error("XLSX does not exist: %s", xlsx_path)
        return 1

    try:
        workbook = load_workbook(xlsx_path, data_only=True)
        tables, issues = build_tables(workbook, reporter)
        if issues:
            reporter.error("validate=FAIL issues=%d", len(issues))
            for issue in issues:
                reporter.error(" - %s", issue)
            return 2
        reporter.info("validate=OK")

        meta = {}
        meta_table = tables.get("Meta")
        if meta_table and meta_table.get("rows"):
            meta = meta_table["rows"][0]
        data = {
            "meta": meta,
            "tables": tables,
        }
        indent = None if args.no_pretty else 2
        json_text = json.dumps(data, ensure_ascii=False, indent=indent)
        json_bytes = json_text.encode("utf-8")
        reporter.info("json_bytes=%d", len(json_bytes))

        if args.validate_only:
            elapsed_ms = (time.perf_counter() - start_time) * 1000
            reporter.success("validate_only elapsed_ms=%.2f", elapsed_ms)
            return 0

        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_bytes(json_bytes)
        size_bytes = out_path.stat().st_size
        elapsed_ms = (time.perf_counter() - start_time) * 1000
        reporter.success("out_bytes=%d elapsed_ms=%.2f", size_bytes, elapsed_ms)
        return 0
    except Exception as exc:
        reporter.error("Unhandled exception during export: %s", exc)
        return 3


def main() -> None:
    sys.exit(run(sys.argv))


if __name__ == "__main__":
    main()
