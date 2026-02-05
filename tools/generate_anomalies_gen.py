#!/usr/bin/env python3
"""Generate 500-day AnomaliesGen data with pacing rules."""

from openpyxl import load_workbook

# Cooldown periods - spawnCount MUST be 0
COOLDOWN_PERIODS = [
    (97, 100),
    (167, 170),
    (239, 242),
    (309, 312),
    (389, 392),
    (459, 462),
]

# Spike weeks - add +3
SPIKE_WEEKS = [
    (90, 96),
    (160, 166),
    (230, 238),
    (300, 308),
    (380, 388),
    (450, 458),
]


def get_base_value(day: int) -> int:
    """Get base spawn count for a day."""
    if 1 <= day <= 80:
        return 1
    elif 81 <= day <= 200:
        return 2
    elif 201 <= day <= 320:
        return 3
    elif 321 <= day <= 420:
        return 4
    else:  # 421-500
        return 5


def get_wave_modifier(day: int) -> int:
    """Get wave modifier based on 20-day cycle."""
    cycle_day = ((day - 1) % 20) + 1
    if 1 <= cycle_day <= 8:
        return 0
    elif 9 <= cycle_day <= 14:
        return 1
    elif 15 <= cycle_day <= 18:
        return 0
    else:  # 19-20
        return -1


def is_in_range(day: int, ranges: list[tuple[int, int]]) -> bool:
    """Check if day is in any of the given ranges."""
    return any(start <= day <= end for start, end in ranges)


def calculate_spawn_count(day: int) -> int:
    """Calculate spawn count for a given day."""
    # Cooldown has highest priority
    if is_in_range(day, COOLDOWN_PERIODS):
        return 0
    
    # Calculate base + wave + spike
    base = get_base_value(day)
    wave = get_wave_modifier(day)
    spike = 3 if is_in_range(day, SPIKE_WEEKS) else 0
    
    # Clamp to [0, 8]
    result = base + wave + spike
    return max(0, min(8, result))


def generate_data():
    """Generate 500 days of data."""
    data = []
    for day in range(1, 501):
        spawn_count = calculate_spawn_count(day)
        data.append((day, spawn_count))
    return data


def update_excel(xlsx_path: str):
    """Update the AnomaliesGen sheet with generated data."""
    print(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    
    if 'AnomaliesGen' not in wb.sheetnames:
        raise ValueError("AnomaliesGen sheet not found!")
    
    ws = wb['AnomaliesGen']
    
    # Generate data
    print("Generating 500-day data...")
    data = generate_data()
    
    # Clear existing data rows (keep headers in rows 1-3)
    print("Clearing existing data rows...")
    ws.delete_rows(4, ws.max_row - 3)
    
    # Write new data
    print("Writing new data...")
    for day, spawn_count in data:
        ws.append([day, spawn_count])
    
    # Save workbook
    print(f"Saving workbook...")
    wb.save(xlsx_path)
    print("Done!")
    
    return data


def calculate_statistics(data):
    """Calculate and print statistics."""
    print("\n=== Statistics ===")
    
    # Average spawn count per 50-day segment
    print("\nAverage spawn count per 50-day segment:")
    for i in range(0, 500, 50):
        segment = data[i:i+50]
        avg = sum(spawn for _, spawn in segment) / len(segment)
        print(f"  Days {i+1}-{i+50}: {avg:.2f}")
    
    # Peak spawn counts during spike weeks
    print("\nPeak spawn counts during spike weeks:")
    for start, end in SPIKE_WEEKS:
        segment = [(day, spawn) for day, spawn in data if start <= day <= end]
        max_spawn = max(spawn for _, spawn in segment)
        print(f"  Days {start}-{end}: {max_spawn}")
    
    # Verify cooldown periods are all 0
    print("\nCooldown period verification:")
    all_zero = True
    for start, end in COOLDOWN_PERIODS:
        segment = [(day, spawn) for day, spawn in data if start <= day <= end]
        non_zero = [day for day, spawn in segment if spawn != 0]
        if non_zero:
            print(f"  Days {start}-{end}: ERROR - Non-zero days: {non_zero}")
            all_zero = False
        else:
            print(f"  Days {start}-{end}: OK (all 0)")
    
    if all_zero:
        print("\n✓ All cooldown periods verified as 0")
    else:
        print("\n✗ Some cooldown periods have non-zero values!")
    
    # Overall statistics
    total_spawns = sum(spawn for _, spawn in data)
    avg_overall = total_spawns / len(data)
    max_spawn = max(spawn for _, spawn in data)
    print(f"\nOverall statistics:")
    print(f"  Total spawns: {total_spawns}")
    print(f"  Average per day: {avg_overall:.2f}")
    print(f"  Maximum spawn count: {max_spawn}")


if __name__ == "__main__":
    import sys
    
    xlsx_path = "GameData/Local/game_data.xlsx"
    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]
    
    data = update_excel(xlsx_path)
    calculate_statistics(data)
